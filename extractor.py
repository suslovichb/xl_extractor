import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from tkinter import Tk, Text, StringVar, BooleanVar, _setit, messagebox, filedialog

class ExtractorApp(tk.Frame):
    def __init__(self, parent, *args, **kwargs):
        tk.Frame.__init__(self, parent, *args, **kwargs)
        self.parent = parent
        self.parent.title('xl')
        self.parent.title('XL Extractor')
        self.parent.geometry('510x582')
        self.parent.resizable(0, 0)

        self.source_workbook = None
        self.extraction_workbook = None

        menu_button_style = ttk.Style()
        menu_button_style.configure("TMenubutton", foreground="#ffffff00", background="gray83", font=('Arial', 6))
        menu_button_style.map("TMenubutton", foreground=[("disabled","#FFFFFF00")])
        secondary_button_style = ttk.Style()
        secondary_button_style.configure('secondary.TButton', font=('Calibri', 8))

        self.source_workbook_label = tk.Label(self, text="Source workbook")
        self.source_workbook_field = tk.Text(self, width=50, height = 1, state = 'disabled')
        self.open_source_file_button = ttk.Button(self, text="Open", command=self.choose_source_file, style='secondary.TButton')
        self.source_worksheet_label = tk.Label(self, text="Source worksheet")
        self.source_worksheet_field = tk.Text(self, width=50, height = 1, state = 'disabled')
        self.source_worksheet_choice = tk.StringVar()
        self.source_worksheet_choice.set("")
        self.source_worksheet_choice.trace("w", self.source_worksheet_changed)
        self.source_worksheet_optionmenu = ttk.OptionMenu(self, self.source_worksheet_choice)
        self.extraction_workbook_label = tk.Label(self, text="Extraction data workbook")
        self.extraction_workbook_field = tk.Text(self, width=50, height = 1, state = 'disabled')
        self.open_extraction_file_button = ttk.Button(self, text="Open", command=self.choose_extraction_file, style='secondary.TButton')
        self.extraction_worksheet_label = tk.Label(self, text="Extraction data worksheet")
        self.extraction_worksheet_field = tk.Text(self, width=50, height = 1, state = 'disabled')
        self.extraction_worksheet_choice = tk.StringVar()
        self.extraction_worksheet_choice.set("")
        self.extraction_worksheet_choice.trace("w", self.extraction_worksheet_changed)
        self.extraction_worksheet_optionmenu = ttk.OptionMenu(self, self.extraction_worksheet_choice)
        self.extraction_columns_frame = tk.Frame(self)
        self.extraction_columns_label = tk.Label(self, text="Extraction data columns")
        self.extraction_columns_listbox = tk.Listbox(self.extraction_columns_frame, width=24, height=10, selectmode='multiple')
        self.extraction_columns_scroll = tk.Scrollbar(self.extraction_columns_frame)
        self.extraction_columns_listbox.config(yscrollcommand=self.extraction_columns_scroll.set)
        self.extraction_columns_scroll.config(command=self.extraction_columns_listbox.yview)
        self.extract_button = ttk.Button(self, text="EXTRACT", command=self.extract)
        self.progress = ttk.Progressbar(self, orient='horizontal', length=500, mode = 'determinate')
        self.status_text = tk.StringVar()
        self.status_label = tk.Label(self, text='', font=('Arial', 8), fg='green', textvariable=self.status_text)

        self.source_workbook_label.grid(row=0, column=0, columnspan=5, pady=(20,5))
        self.source_workbook_field.grid(row=1, column=0, columnspan=3, padx=(10,5))
        self.open_source_file_button.grid(row=1, column=4, sticky="wns")
        self.source_worksheet_label.grid(row=2, column=0, columnspan=5, pady=(10,5))
        self.source_worksheet_field.grid(row=3, column=0, columnspan=3, padx=(10,5))
        self.source_worksheet_optionmenu.grid(row=3, column=4, sticky="wns")
        self.extraction_workbook_label.grid(row=4, column=0, columnspan=5, pady=(30,5))
        self.extraction_workbook_field.grid(row=5, column=0, columnspan=3, padx=(10,5))
        self.open_extraction_file_button.grid(row=5, column=4, sticky="wns")
        self.extraction_worksheet_label.grid(row=6, column=0, columnspan=5, pady=(10,5))
        self.extraction_worksheet_field.grid(row=7, column=0, columnspan=3, padx=(10,5))
        self.extraction_worksheet_optionmenu.grid(row=7, column=4, sticky="wns")
        self.extraction_columns_label.grid(row=8, column=0, columnspan=5, pady=(10,5))
        self.extraction_columns_frame.grid(row=9, column=0, columnspan=5)
        self.extraction_columns_listbox.pack(side='left', fill='both', expand=1)
        self.extraction_columns_scroll.pack(side='right', fill='y')
        self.extract_button.grid(row=10, column=0, columnspan=5, pady=(40,10))
        self.status_label.grid(row=11, column=0, columnspan=5)
        self.progress.grid(row=12, column=0, columnspan=5, padx=5)
        self.pack(side="top", fill="both", expand=True)

    def set_status(self, text):
        self.status_text.set(str(text))
        self.status_label.update()

    def set_progress(self, num):
        self.progress['value'] = num
        self.progress.update()

    def disable_buttons(self):
        self.open_source_file_button['state'] = 'disabled'
        self.source_worksheet_optionmenu['state'] = 'disabled'
        self.open_extraction_file_button['state'] = 'disabled'
        self.extraction_worksheet_optionmenu['state'] = 'disabled'
        self.extract_button['state'] = 'disabled'

    def enable_buttons(self):
        self.open_source_file_button['state'] = 'normal'
        self.source_worksheet_optionmenu['state'] = 'normal'
        self.open_extraction_file_button['state'] = 'normal'
        self.extraction_worksheet_optionmenu['state'] = 'normal'
        self.extract_button['state'] = 'normal'

    def _lock_buttons(func):
        def wrapper(self):
            self.disable_buttons()
            func(self)
            self.enable_buttons()
        return wrapper

    def _ensure_empty_progress(func):
        def wrapper(self):
            self.set_progress(0)
            func(self)
            self.set_progress(0)
        return wrapper

    def _ensure_empty_status(func):
        def wrapper(self):
            self.set_status('')
            func(self)
            self.set_status('')
        return wrapper

    def get_workbook(self, file_path, error_message='Workbook loading failed'):
        try:
            return load_workbook(file_path)
        except Exception:
            messagebox.showerror(title="Error", message=error_message)

    @_lock_buttons
    def choose_source_file(self):
        source_file = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsx"),("All files","*.*")))
        if source_file:
            self.source_workbook_field.configure(state = 'normal')
            self.source_workbook_field.delete(1.0, 'end')
            self.source_workbook_field.insert(1.0, source_file)
            self.source_workbook_field.configure(state = 'disabled')
            self.set_status('Source workbook loading...')
            self.source_workbook = self.get_workbook(source_file)
            self.set_status('')
            self.refresh_source_worksheets()

    def refresh_source_worksheets(self):
        self.source_worksheet_choice.set("")
        self.source_worksheet_optionmenu['menu'].delete(0, 'end')
        new_choices = self.source_workbook.sheetnames
        for choice in new_choices:
            self.source_worksheet_optionmenu['menu'].add_command(label=choice, command=_setit(self.source_worksheet_choice, choice))

    def source_worksheet_changed(self, *args):
        self.source_worksheet_field.configure(state = 'normal')
        self.source_worksheet_field.delete(1.0, 'end')
        self.source_worksheet_field.insert(1.0, self.source_worksheet_choice.get())
        self.source_worksheet_field.configure(state = 'disabled')

    @_lock_buttons
    def choose_extraction_file(self):
        extraction_file = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsx"),("All files","*.*")))
        if extraction_file:
            self.extraction_workbook_field.configure(state = 'normal')
            self.extraction_workbook_field.delete(1.0, 'end')
            self.extraction_workbook_field.insert(1.0, extraction_file)
            self.extraction_workbook_field.configure(state = 'disabled')
            self.set_status('Extraction workbook loading...')
            self.extraction_workbook = self.get_workbook(extraction_file)
            self.set_status('')
            self.refresh_extraction_worksheets()
            self.clear_extraction_columns_listbox()

    def refresh_extraction_worksheets(self):
        self.extraction_worksheet_choice.set("")
        self.extraction_worksheet_optionmenu['menu'].delete(0, 'end')
        new_choices = self.extraction_workbook.sheetnames
        for choice in new_choices:
            self.extraction_worksheet_optionmenu['menu'].add_command(label=choice, command=_setit(self.extraction_worksheet_choice, choice))

    def extraction_worksheet_changed(self, *args):
        self.extraction_worksheet_field.configure(state = 'normal')
        self.extraction_worksheet_field.delete(1.0, 'end')
        self.extraction_worksheet_field.insert(1.0, self.extraction_worksheet_choice.get())
        self.extraction_worksheet_field.configure(state = 'disabled')
        self.refresh_extraction_columns()

    def refresh_extraction_columns(self):
        self.clear_extraction_columns_listbox()
        if self.extraction_worksheet_choice.get():
            try:
                for header_cell in self.extraction_workbook[self.extraction_worksheet_choice.get()][1]:
                    self.extraction_columns_listbox.insert('end', header_cell.value)
            except:
                messagebox.showerror(title="Error", message='Columns loading failed')

    def clear_extraction_columns_listbox(self):
        self.extraction_columns_listbox.delete(0, 'end')

    def find_col_index(self, ws, col_name):
        for cell in ws[1]:
            if cell.value == col_name:
                return cell.col_idx

    @_ensure_empty_progress
    @_ensure_empty_status
    @_lock_buttons
    def extract(self):
        self.progress["maximum"] = 8
        self.set_progress(0)

        source_cols_indices_by_names = {}
        extraction_cols_indices_by_names = {}
        check_list = []
        rows_to_extract = []

        self.set_status('Source sheet reading...')

        source_worksheet_name = self.source_worksheet_field.get(1.0, 'end').replace('\n','')
        try:
            source_worksheet = self.source_workbook[source_worksheet_name]
        except Exception:
            messagebox.showerror(title="Error", message="Source sheet loading error")
            return

        self.set_progress(1)
        self.set_status('Extraction sheet reading...')

        extraction_worksheet_name = self.extraction_worksheet_field.get(1.0, 'end').replace('\n','')
        try:
            extraction_worksheet = self.extraction_workbook[extraction_worksheet_name]
        except Exception:
            messagebox.showerror(title="Error", message="Extraction sheet loading error")
            return

        self.set_progress(2)
        self.set_status('Selected columns reading...')

        extraction_cols_names = [self.extraction_columns_listbox.get(index) for index in self.extraction_columns_listbox.curselection()]
        for col_name in extraction_cols_names:
            extraction_col_index = self.find_col_index(extraction_worksheet, col_name)
            if not extraction_col_index:
                messagebox.showerror(title="Error", message="Column {} not found in extraction worksheet".format(col_name))
                return
            extraction_cols_indices_by_names[col_name] = extraction_col_index
            source_col_index = self.find_col_index(source_worksheet, col_name)
            if not source_col_index:
                messagebox.showerror(title="Error", message="Column {} not found source worksheet".format(col_name))
                return
            source_cols_indices_by_names[col_name] = source_col_index

        self.set_progress(3)
        self.set_status('Check list creating...')

        for row in extraction_worksheet.iter_rows(min_row=2):
            check_list.append({col_name: row[col_index-1].value for col_name, col_index in extraction_cols_indices_by_names.items()})

        self.set_progress(4)
        self.set_status('Searching for rows to extract...')

        for row in source_worksheet.iter_rows(min_row=2):
            dict_to_check = {col_name: row[col_index-1].value for col_name, col_index in source_cols_indices_by_names.items()}
            if dict_to_check in check_list:
                rows_to_extract.append(row[0].row)

        self.set_progress(5)
        self.set_status('Extracted rows saving...')

        extracted_worksheet = self.source_workbook.copy_worksheet(source_worksheet)
        extracted_worksheet.title = "Extracted"
        self.delete_rows_by_list(extracted_worksheet, [i for i in range(2, source_worksheet.max_row+1) if i not in rows_to_extract])

        self.set_progress(6)
        self.set_status('Remainder rows saving...')

        remainder_worksheet = self.source_workbook.copy_worksheet(source_worksheet)
        remainder_worksheet.title = "Remainder"
        self.delete_rows_by_list(remainder_worksheet, rows_to_extract)

        self.set_progress(7)
        self.set_status('Final file saving...')

        try:
            self.save_workbook(self.source_workbook)
        except PermissionError:
            messagebox.showerror(title="Error", message='Permission error')
        else:
            self.set_progress(8)
            self.set_status('Extraction completed!')
            print("Done!")
            messagebox.showinfo(title="Extraction", message="Extraction completed!")
            return

    def delete_rows_by_list(self, worksheet, rows_list):
        rows_sequences = self.get_sequences(rows_list)
        for sequence in reversed(rows_sequences):
            worksheet.delete_rows(sequence[0], sequence[1])

    def get_sequences(self, list_of_ints):
        sequence_count = 1
        sequences = []
        for item in list_of_ints:
            next_item = None
            if list_of_ints.index(item) < (len(list_of_ints) - 1):
                next_item = list_of_ints[list_of_ints.index(item) + 1]

            if (item + 1) == next_item:
                sequence_count += 1
            else:
                first_in_sequence = list_of_ints[list_of_ints.index(item) - sequence_count + 1]
                sequences.append([first_in_sequence, sequence_count])
                sequence_count = 1
        return sequences

    def save_workbook(self, wb):
        filename = self.source_workbook_field.get(1.0, 'end').replace('\n','')
        wb.save(filename)


def main():
    root = tk.Tk()
    app = ExtractorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
